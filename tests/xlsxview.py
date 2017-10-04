from __future__ import absolute_import, print_function, division
import locale
from petl.util.base import Table


class XLSXView(Table):

    def __init__(self, filename, sheet=None, range_string=None,
                 row_offset=0, column_offset=0, **kwargs):
        self.filename = filename
        self.sheet = sheet
        self.range_string = range_string
        self.row_offset = row_offset
        self.column_offset = column_offset
        self.kwargs = kwargs

    def __iter__(self):
        import openpyxl
        wb = openpyxl.load_workbook(filename=self.filename,
                                    read_only=True, **self.kwargs)
        if self.sheet is None:
            ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        elif isinstance(self.sheet, int):
            ws = wb.get_sheet_by_name(wb.get_sheet_names()[self.sheet])
        else:
            ws = wb.get_sheet_by_name(str(self.sheet))

        for row in ws.iter_rows(range_string=self.range_string,
                                row_offset=self.row_offset,
                                column_offset=self.column_offset):
            yield tuple(cell.value for cell in row)
        try:
            wb._archive.close()
        except AttributeError as e:
            # just here in case openpyxl stops exposing an _archive property.
            pass

def fromxlsx(filename, sheet=None, range_string=None, row_offset=0, column_offset=0, **kwargs):
    return XLSXView(filename, sheet=sheet, range_string=range_string,
                    row_offset=row_offset, column_offset=column_offset,
                    **kwargs)

def toxlsx(tbl, filename, sheet=None, encoding=None):
    import openpyxl
    if encoding is None:
        encoding = locale.getpreferredencoding()
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet)
    for row in tbl:
        ws.append(row)
    wb.save(filename)

#Table.toxlsx = toxlsx
