from openpyxl import load_workbook, Workbook
#pip install -I openpyxl==2.3.3
filepath = "./demo.xlsx"
wb = Workbook()
wb.save(filepath)

wb = load_workbook('test.xlsx')
ws = wb['Sheet1']
ws['D1'] = 'CHANGE D1'
ws['D2'] = 'DATA'
ws['D3'] = 'Mi data'

wb.save('test.xlsx')


"""
wb2 = load_workbook('test.xlsx')
Sheet = wb2.get_sheet_by_name('Sheet1')

print(Sheet.columns)
for i, cellObj in enumerate(Sheet.columns[2], 1):
    cellObj.value = '=IF($A${0}=$B${0}, "COINCIDE", "PASS")'.format(i)
    #pass
    #next(sheet.columns)
wb2.save('test.xlsx') #"""
